from flask import Flask, render_template, request, redirect, url_for, send_file, session
from functools import wraps
import sqlite3
from openpyxl import Workbook
import tempfile
from datetime import date

app = Flask(__name__)
import os
DB_NAME = os.environ.get("KJ_DB_PATH", "database.db")


# -------------------------
# ZÁKLADNÍ FIREMNÍ LOGIN
# -------------------------
# ZMĚŇ SI OBOJE:
app.secret_key = "ZMEN_TO_NA_NECO_DLOUHEHO_TAJNEHO_123456789"
APP_USERNAME = "firma"
APP_PASSWORD = "1234"


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


# =========================
# DATABASE
# =========================
def get_db():
    conn = sqlite3.connect(DB_NAME)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS ridici (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            jmeno TEXT UNIQUE
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS auta (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            spz TEXT UNIQUE,
            popis TEXT
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS jizdy (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            datum TEXT,
            ridic TEXT,
            vozidlo TEXT,
            ucel TEXT,
            km_start INTEGER,
            km_konec INTEGER,
            km_celkem INTEGER
        )
    """)

    conn.commit()
    conn.close()


# =========================
# AUTH ROUTY
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    error = None

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "").strip()

        if username == APP_USERNAME and password == APP_PASSWORD:
            session["logged_in"] = True
            session["username"] = username
            return redirect(url_for("index"))
        else:
            error = "Špatné jméno nebo heslo."

    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


# =========================
# APLIKACE (chráněné routy)
# =========================
@app.route("/")
@login_required
def index():
    conn = get_db()
    jizdy = conn.execute("SELECT * FROM jizdy ORDER BY datum DESC").fetchall()
    conn.close()
    return render_template("index.html", jizdy=jizdy)


@app.route("/add", methods=["GET", "POST"])
@login_required
def add():
    conn = get_db()

    if request.method == "POST":
        # načtení a kontrola km
        try:
            km_start = int(request.form["km_start"])
            km_konec = int(request.form["km_konec"])
        except ValueError:
            conn.close()
            return "Chyba: KM musí být číslo", 400

        km_celkem = km_konec - km_start
        if km_celkem < 0:
            conn.close()
            return "Chyba: KM konec nesmí být menší než KM start", 400

        conn.execute("""
            INSERT INTO jizdy (datum, ridic, vozidlo, ucel, km_start, km_konec, km_celkem)
            VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            request.form["datum"],
            request.form["ridic"],
            request.form["vozidlo"],
            request.form["ucel"],
            km_start,
            km_konec,
            km_celkem
        ))
        conn.commit()
        conn.close()
        return redirect(url_for("index"))

    ridici = conn.execute("SELECT jmeno FROM ridici ORDER BY jmeno").fetchall()
    auta = conn.execute("SELECT spz FROM auta ORDER BY spz").fetchall()
    conn.close()
    return render_template("add.html", ridici=ridici, auta=auta)


@app.route("/ridici", methods=["GET", "POST"])
@login_required
def ridici():
    conn = get_db()

    if request.method == "POST":
        jmeno = request.form.get("jmeno", "").strip()
        if jmeno:
            try:
                conn.execute("INSERT INTO ridici (jmeno) VALUES (?)", (jmeno,))
                conn.commit()
            except:
                # duplicitní jméno apod.
                pass

    ridici_list = conn.execute("SELECT * FROM ridici ORDER BY jmeno").fetchall()
    conn.close()
    return render_template("ridici.html", ridici=ridici_list)


@app.route("/auta", methods=["GET", "POST"])
@login_required
def auta():
    conn = get_db()

    if request.method == "POST":
        spz = request.form.get("spz", "").strip()
        popis = request.form.get("popis", "").strip()
        if spz:
            try:
                conn.execute("INSERT INTO auta (spz, popis) VALUES (?, ?)", (spz, popis))
                conn.commit()
            except:
                # duplicitní spz apod.
                pass

    auta_list = conn.execute("SELECT * FROM auta ORDER BY spz").fetchall()
    conn.close()
    return render_template("auta.html", auta=auta_list)


# ---------- EXPORT FORM (měsíc) ----------
@app.route("/export_mesic", methods=["GET", "POST"])
@login_required
def export_mesic():
    if request.method == "POST":
        mesic = request.form["mesic"]  # YYYY-MM
        return redirect(url_for("export_spz_mesic", mesic=mesic))

    return render_template("export_mesic.html", now=date.today().strftime("%Y-%m"))


# ---------- EXPORT EXCEL (SPZ sheets + měsíční součet) ----------
@app.route("/export_spz_mesic/<mesic>")
@login_required
def export_spz_mesic(mesic):
    conn = get_db()

    auta_list = conn.execute("SELECT spz FROM auta ORDER BY spz").fetchall()
    jizdy = conn.execute("""
        SELECT * FROM jizdy
        WHERE substr(datum, 1, 7) = ?
        ORDER BY datum
    """, (mesic,)).fetchall()

    conn.close()

    wb = Workbook()

    for auto in auta_list:
        spz = auto["spz"]
        ws = wb.create_sheet(title=spz)

        ws.append(["Datum", "Řidič", "Účel", "KM start", "KM konec", "KM celkem"])

        total_km = 0
        for j in jizdy:
            if j["vozidlo"] == spz:
                ws.append([
                    j["datum"],
                    j["ridic"],
                    j["ucel"],
                    j["km_start"],
                    j["km_konec"],
                    j["km_celkem"]
                ])
                total_km += int(j["km_celkem"])

        ws.append([])
        ws.append(["", "", "SOUČET KM ZA MĚSÍC", "", "", total_km])

    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)

    return send_file(
        tmp.name,
        as_attachment=True,
        download_name=f"kniha_jizd_{mesic}.xlsx"
    )


# =========================
# START
# =========================
if __name__ == "__main__":
    init_db()
    app.run(host="192.168.0.162", port=5000, debug=False)

